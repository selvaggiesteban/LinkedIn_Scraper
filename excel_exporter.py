"""
excel_exporter.py — Unified Excel + CSV exporter for LinkedIn Scraper.

Produces:
  - output/LinkedIn_Scraper_<ts>.xlsx  (1 workbook, 11 sheets)
  - output/LinkedIn_Scraper_<ts>.csv   (1 flat CSV with all categories)

The 11 sheets are:
  0. README       — guide + legend
  1. jobs
  2. job_details
  3. people
  4. person_profiles
  5. posts_feed
  6. posts_companies
  7. company_search
  8. company_profiles
  9. company_employees
  10. authors

Dependency: openpyxl (pure Python, no binary deps).
"""
from __future__ import annotations

import csv
import json
from datetime import datetime
from pathlib import Path
from typing import Any

try:
    from openpyxl import Workbook
    from openpyxl.styles import Alignment, Font, PatternFill
    from openpyxl.utils import get_column_letter

    _HAVE_OPENPYXL = True
except ImportError:
    _HAVE_OPENPYXL = False


CATEGORIES = [
    "jobs",
    "job_details",
    "people",
    "person_profiles",
    "posts_feed",
    "posts_companies",
    "company_search",
    "company_profiles",
    "company_employees",
    "authors",
]


def _ts() -> str:
    return datetime.now().strftime("%Y%m%d_%H%M%S")


def _flatten(item: dict[str, Any]) -> dict[str, Any]:
    """Normalize an item dict into the unified schema columns."""
    flat = {
        "type": item.get("type", ""),
        "source": item.get("source", ""),
        "search_keyword": item.get("search_keyword", ""),
        "title": item.get("title", "") or item.get("name", ""),
        "company": item.get("company", "") or item.get("company_name", ""),
        "location": item.get("location", ""),
        "url": (
            item.get("url")
            or item.get("job_url")
            or item.get("profile_url")
            or item.get("post_url")
            or item.get("applyUrl")
            or ""
        ),
        "external_id": item.get("external_id", "") or str(item.get("jobId", "") or item.get("username", "") or item.get("slug", "") or ""),
        "posted_date": item.get("posted_date", "") or item.get("postedDate", ""),
        "scraped_at": item.get("scraped_at", ""),
        "text_ocr": item.get("text_ocr", ""),
        "is_valid": "",
    }
    v = item.get("validation")
    if isinstance(v, dict):
        flat["is_valid"] = str(v.get("is_valid", ""))
    return flat


def _column_order(category: str) -> list[str]:
    base = ["type", "source", "search_keyword", "title", "company", "location", "url", "external_id", "posted_date", "scraped_at", "text_ocr", "is_valid"]
    if category == "job_details":
        base = ["type", "source", "title", "company", "location", "url", "external_id", "Description", "posted_date", "scraped_at", "text_ocr", "is_valid"]
    elif category == "person_profiles":
        base = ["type", "source", "name", "headline", "location", "url", "external_id", "experience", "education", "about", "scraped_at", "text_ocr"]
    elif category == "people":
        base = ["type", "source", "name", "headline", "location", "url", "external_id", "scraped_at"]
    elif category == "company_profiles":
        base = ["type", "source", "name", "tagline", "industry", "size", "location", "url", "external_id", "scraped_at", "text_ocr"]
    elif category == "posts_feed":
        base = ["type", "source", "author", "text", "url", "posted_date", "scraped_at"]
    elif category == "posts_companies":
        base = ["type", "source", "author", "company_name", "text", "url", "posted_date", "scraped_at"]
    elif category == "company_search":
        base = ["type", "source", "name", "url", "external_id", "scraped_at"]
    elif category == "company_employees":
        base = ["type", "source", "name", "role", "connection_degree", "location", "url", "external_id", "scraped_at"]
    elif category == "authors":
        base = ["type", "source", "name", "url", "post_count", "last_active", "scraped_at"]
    return base


def _get_cell_value(item: dict[str, Any], col: str) -> Any:
    direct = item.get(col)
    if direct is not None:
        return direct
    aliases = {
        "title": ("name",),
        "name": ("title",),
        "company": ("company_name",),
        "Description": ("description", "text_ocr", "text"),
        "experience": ("experience_text",),
        "education": ("education_text",),
        "about": ("about_text", "text_ocr", "text"),
        "author": ("author_name", "name"),
        "text": ("text_ocr",),
        "url": ("job_url", "profile_url", "post_url", "applyUrl"),
        "external_id": ("jobId", "username", "slug"),
        "posted_date": ("postedDate",),
    }
    for alias in aliases.get(col, ()):
        v = item.get(alias)
        if v is not None:
            return v
    return ""


def _write_sheet(ws, items: list[dict[str, Any]], category: str) -> int:
    cols = _column_order(category)
    header_fill = PatternFill("solid", fgColor="1F4E78")
    header_font = Font(color="FFFFFF", bold=True)
    for c_idx, col in enumerate(cols, 1):
        cell = ws.cell(row=1, column=c_idx, value=col)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal="center", vertical="center")
        ws.column_dimensions[get_column_letter(c_idx)].width = max(12, min(40, len(col) + 5))
    ws.freeze_panes = "A2"
    written = 0
    for r_idx, item in enumerate(items, 2):
        for c_idx, col in enumerate(cols, 1):
            value = _get_cell_value(item, col)
            if isinstance(value, (dict, list)):
                value = json.dumps(value, ensure_ascii=False)
            elif isinstance(value, str) and len(value) > 32000:
                value = value[:32000] + "…[truncated]"
            ws.cell(row=r_idx, column=c_idx, value=value)
        written += 1
    return written


def _write_readme_sheet(wb) -> None:
    ws = wb.create_sheet("README", 0)
    title_font = Font(size=16, bold=True, color="1F4E78")
    header_font = Font(size=12, bold=True)
    ws.column_dimensions["A"].width = 30
    ws.column_dimensions["B"].width = 80
    ws.cell(row=1, column=1, value="LinkedIn Scraper — Export Workbook").font = title_font
    rows = [
        ("", ""),
        ("Generated", datetime.now().isoformat()),
        ("Total categories", str(len(CATEGORIES))),
        ("", ""),
        ("Sheet", "Contents"),
        ("jobs", "Job listings from Guest API + MCP search_jobs"),
        ("job_details", "Full job descriptions (MCP get_job_details + Scrapling)"),
        ("people", "Recruiter / HR profiles discovered via MCP search_people"),
        ("person_profiles", "Detailed person profiles (MCP get_person_profile + Scrapling)"),
        ("posts_feed", "Personal feed posts (MCP get_feed)"),
        ("posts_companies", "Company page posts (MCP get_company_posts + Scrapling)"),
        ("company_search", "Companies discovered via MCP search_companies"),
        ("company_profiles", "Company /about pages (Guest API + MCP + Scrapling)"),
        ("company_employees", "Employees of discovered companies (MCP get_company_employees)"),
        ("authors", "Authors deduced from posts_feed + posts_companies"),
        ("", ""),
        ("Legend", ""),
        ("text_ocr", "Free text extracted via EasyOCR from a screenshot of the page (Spanish+English)"),
        ("is_valid", "True if both primary keyword AND (secondary keyword OR hashtag) appear in the text"),
        ("source", "guest_api | mcp | scrapling | ocr"),
        ("external_id", "LinkedIn jobId for jobs, username for people, slug for companies"),
    ]
    r = 3
    for label, value in rows:
        ws.cell(row=r, column=1, value=label).font = header_font if label and not value and label in ("Sheet", "Legend") else Font()
        ws.cell(row=r, column=2, value=value)
        r += 1


def _export_xlsx(results: dict[str, list[dict[str, Any]]], output_dir: Path) -> Path:
    if not _HAVE_OPENPYXL:
        raise ImportError("openpyxl not installed — pip install openpyxl")
    wb = Workbook()
    # Remove default sheet
    if "Sheet" in wb.sheetnames:
        wb.remove(wb["Sheet"])
    # README sheet at index 0
    _write_readme_sheet(wb)
    # One sheet per category
    counts: dict[str, int] = {}
    for cat in CATEGORIES:
        ws = wb.create_sheet(cat)
        items = results.get(cat, [])
        counts[cat] = _write_sheet(ws, items, cat)
    path = output_dir / f"LinkedIn_Scraper_{_ts()}.xlsx"
    wb.save(path)
    return path


def _export_csv_flat(results: dict[str, list[dict[str, Any]]], output_dir: Path) -> Path:
    path = output_dir / f"LinkedIn_Scraper_{_ts()}.csv"
    cols = ["category", "type", "source", "search_keyword", "title", "company", "location", "url", "external_id", "posted_date", "scraped_at", "text_ocr", "is_valid"]
    with open(path, "w", encoding="utf-8-sig", newline="") as f:
        writer = csv.DictWriter(f, fieldnames=cols, extrasaction="ignore")
        writer.writeheader()
        for cat in CATEGORIES:
            for item in results.get(cat, []):
                flat = _flatten(item)
                flat["category"] = cat
                writer.writerow(flat)
    return path


def export_all(results: dict[str, list[dict[str, Any]]], output_dir: Path) -> tuple[Path, Path]:
    """Generate unified .xlsx (11 sheets) and flat .csv. Returns (xlsx_path, csv_path)."""
    output_dir.mkdir(parents=True, exist_ok=True)
    if _HAVE_OPENPYXL:
        xlsx_path = _export_xlsx(results, output_dir)
    else:
        xlsx_path = None
    csv_path = _export_csv_flat(results, output_dir)
    if xlsx_path is None:
        print("[excel_exporter] openpyxl missing — skipped .xlsx, only .csv generated")
    return xlsx_path, csv_path
