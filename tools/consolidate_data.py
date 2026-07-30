"""
consolidate_data.py — Consolidate all LinkedIn Scraper historical data.

Reads all unique data files, normalizes to a unified schema, deduplicates
across files, and generates:
  - data/outputs/consolidated/linkedin_all_jobs.json
  - data/outputs/consolidated/linkedin_all_jobs.csv
  - data/outputs/consolidated/linkedin_all_jobs.xlsx

Usage:
    python tools/consolidate_data.py
"""
from __future__ import annotations

import csv
import hashlib
import json
import sys
from datetime import datetime
from pathlib import Path
from typing import Any

PROJECT_ROOT = Path(__file__).resolve().parent.parent
HISTORICAL_DIR = PROJECT_ROOT / "data" / "outputs" / "historical"
OUTPUT_DIR = PROJECT_ROOT / "data" / "outputs" / "consolidated"
ROOT_DIR = PROJECT_ROOT


# ── Unified schema ──────────────────────────────────────────────────────

UNIFIED_FIELDS = [
    "id", "type", "source", "search_keyword", "title", "company",
    "location", "url", "external_id", "posted_date", "scraped_at",
    "origin_date", "origin_file", "text_ocr", "is_valid",
]


def _make_id(url: str, job_id: str = "") -> str:
    """Generate a stable dedup key from URL or jobId."""
    key = job_id or url.rstrip("/").split("?")[0]
    return hashlib.md5(key.encode()).hexdigest()[:12]


def _normalize_url(url: str) -> str:
    """Normalize URL for dedup: strip trailing slash and query params."""
    return url.rstrip("/").split("?")[0]


# ── Readers ─────────────────────────────────────────────────────────────

def _read_guest_api_json(path: Path, origin_file: str) -> list[dict]:
    """Read Guest API JSON (Jul 15 or Jul 23 format)."""
    with open(path, "r", encoding="utf-8") as f:
        data = json.load(f)
    records = []
    for item in data:
        url = item.get("job_url", "")
        records.append({
            "id": _make_id(url, item.get("external_id", "")),
            "type": "job",
            "source": item.get("source", "guest_api"),
            "search_keyword": item.get("search_keyword", ""),
            "title": item.get("title", ""),
            "company": item.get("company", ""),
            "location": item.get("location", ""),
            "url": _normalize_url(url),
            "external_id": item.get("external_id", ""),
            "posted_date": item.get("posted_date", ""),
            "scraped_at": item.get("scraped_at", ""),
            "origin_date": path.stem.split("_")[1] if "_" in path.stem else "",
            "origin_file": origin_file,
            "text_ocr": item.get("text_ocr", ""),
            "is_valid": _extract_is_valid(item),
        })
    return records


def _read_mcp_json(path: Path, origin_file: str) -> list[dict]:
    """Read MCP historical JSON (different schema)."""
    with open(path, "r", encoding="utf-8") as f:
        data = json.load(f)
    records = []
    for item in data:
        job_id = str(item.get("jobId", ""))
        url = item.get("applyUrl", "")
        if not url and job_id:
            url = f"https://www.linkedin.com/jobs/view/{job_id}/"
        records.append({
            "id": _make_id(url, job_id),
            "type": "job",
            "source": "mcp",
            "search_keyword": "",
            "title": item.get("title", ""),
            "company": item.get("companyName", ""),
            "location": item.get("location", ""),
            "url": _normalize_url(url),
            "external_id": job_id,
            "posted_date": item.get("postedDate", ""),
            "scraped_at": "",
            "origin_date": "20260714",
            "origin_file": origin_file,
            "text_ocr": "",
            "is_valid": "",
        })
    return records


def _read_empresas_json(path: Path, origin_file: str) -> list[dict]:
    """Read empresas JSON (company contacts, different domain)."""
    with open(path, "r", encoding="utf-8") as f:
        data = json.load(f)
    records = []
    for item in data:
        url = item.get("sitio_web", "")
        records.append({
            "id": _make_id(url, item.get("nombre", "")),
            "type": "company",
            "source": item.get("fuente", "serper"),
            "search_keyword": "",
            "title": item.get("nombre", ""),
            "company": item.get("nombre", ""),
            "location": "",
            "url": _normalize_url(url) if url else "",
            "external_id": "",
            "posted_date": "",
            "scraped_at": item.get("fecha", ""),
            "origin_date": "20260622",
            "origin_file": origin_file,
            "text_ocr": "",
            "is_valid": "",
            # Extra company fields stored in text_ocr as JSON for traceability
            "_emails": item.get("emails", ""),
            "_telefonos": item.get("telefonos", ""),
            "_sector": item.get("sector", ""),
        })
    return records


def _read_all_results_json(path: Path, origin_file: str) -> list[dict]:
    """Read the all_results envelope JSON (contains jobs in results.jobs)."""
    with open(path, "r", encoding="utf-8") as f:
        data = json.load(f)
    jobs = data.get("results", {}).get("jobs", [])
    records = []
    for item in jobs:
        url = item.get("job_url", "") or item.get("url", "")
        records.append({
            "id": _make_id(url, item.get("external_id", "")),
            "type": "job",
            "source": item.get("source", "guest_api"),
            "search_keyword": item.get("search_keyword", ""),
            "title": item.get("title", ""),
            "company": item.get("company", ""),
            "location": item.get("location", ""),
            "url": _normalize_url(url),
            "external_id": item.get("external_id", ""),
            "posted_date": item.get("posted_date", ""),
            "scraped_at": item.get("scraped_at", ""),
            "origin_date": "20260723",
            "origin_file": origin_file,
            "text_ocr": item.get("text_ocr", ""),
            "is_valid": _extract_is_valid(item),
        })
    return records


def _extract_is_valid(item: dict) -> str:
    """Extract is_valid from nested validation dict or flat field."""
    v = item.get("validation")
    if isinstance(v, dict):
        return str(v.get("is_valid", ""))
    return str(item.get("is_valid", ""))


# ── Dedup ───────────────────────────────────────────────────────────────

def _dedup(records: list[dict]) -> list[dict]:
    """Deduplicate by normalized URL, keeping the most enriched version."""
    by_url: dict[str, dict] = {}
    for r in records:
        url = r.get("url", "")
        if not url:
            # No URL: dedup by id
            key = r["id"]
        else:
            key = url
        existing = by_url.get(key)
        if existing is None:
            by_url[key] = r
        else:
            # Keep the one with more data (text_ocr or validation)
            if r.get("text_ocr") and not existing.get("text_ocr"):
                by_url[key] = r
            elif r.get("is_valid") and not existing.get("is_valid"):
                by_url[key] = r
    return list(by_url.values())


# ── Writers ─────────────────────────────────────────────────────────────

def _write_json(records: list[dict], path: Path) -> None:
    with open(path, "w", encoding="utf-8") as f:
        json.dump(records, f, ensure_ascii=False, indent=2, default=str)


def _write_csv(records: list[dict], path: Path) -> None:
    # Separate jobs and companies for proper column handling
    jobs = [r for r in records if r["type"] == "job"]
    companies = [r for r in records if r["type"] == "company"]

    fieldnames = UNIFIED_FIELDS + ["_emails", "_telefonos", "_sector"]
    with open(path, "w", encoding="utf-8-sig", newline="") as f:
        writer = csv.DictWriter(f, fieldnames=fieldnames, extrasaction="ignore")
        writer.writeheader()
        for r in jobs:
            writer.writerow(r)
        for r in companies:
            writer.writerow(r)


def _write_xlsx(records: list[dict], path: Path) -> None:
    try:
        from openpyxl import Workbook
        from openpyxl.styles import Alignment, Font, PatternFill
        from openpyxl.utils import get_column_letter
    except ImportError:
        print("  [WARN] openpyxl not installed — skipping .xlsx")
        return

    jobs = [r for r in records if r["type"] == "job"]
    companies = [r for r in records if r["type"] == "company"]

    wb = Workbook()

    # ── README sheet ──
    ws_readme = wb.active
    ws_readme.title = "README"
    title_font = Font(size=16, bold=True, color="1F4E78")
    header_font = Font(size=12, bold=True)
    ws_readme.column_dimensions["A"].width = 30
    ws_readme.column_dimensions["B"].width = 80
    ws_readme.cell(row=1, column=1, value="LinkedIn Scraper — Consolidated Dataset").font = title_font
    rows_info = [
        ("", ""),
        ("Generated", datetime.now().isoformat()),
        ("Total jobs", str(len(jobs))),
        ("Total companies", str(len(companies))),
        ("", ""),
        ("Sheet", "Contents"),
        ("jobs", "All unique job listings from Guest API + MCP + JS Scraper"),
        ("companies", "Company contacts from Serper API / web scraping"),
        ("", ""),
        ("Column", "Description"),
        ("id", "Unique hash for dedup (MD5 of URL or jobId)"),
        ("type", "job | company"),
        ("source", "guest_api | mcp | js_scraper | serper"),
        ("origin_date", "Date of the original scraping run (YYYYMMDD)"),
        ("origin_file", "Source filename for traceability"),
    ]
    r = 3
    for label, value in rows_info:
        ws_readme.cell(row=r, column=1, value=label).font = header_font if label and not value and label in ("Sheet", "Column") else Font()
        ws_readme.cell(row=r, column=2, value=value)
        r += 1

    # ── Jobs sheet ──
    ws_jobs = wb.create_sheet("jobs")
    job_cols = ["id", "type", "source", "search_keyword", "title", "company",
                "location", "url", "external_id", "posted_date", "scraped_at",
                "origin_date", "origin_file", "text_ocr", "is_valid"]
    header_fill = PatternFill("solid", fgColor="1F4E78")
    header_font_white = Font(color="FFFFFF", bold=True)
    for c_idx, col in enumerate(job_cols, 1):
        cell = ws_jobs.cell(row=1, column=c_idx, value=col)
        cell.fill = header_fill
        cell.font = header_font_white
        cell.alignment = Alignment(horizontal="center")
        ws_jobs.column_dimensions[get_column_letter(c_idx)].width = max(12, min(40, len(col) + 5))
    ws_jobs.freeze_panes = "A2"
    for r_idx, job in enumerate(jobs, 2):
        for c_idx, col in enumerate(job_cols, 1):
            value = job.get(col, "")
            if isinstance(value, (dict, list)):
                value = json.dumps(value, ensure_ascii=False)
            elif isinstance(value, str) and len(value) > 32000:
                value = value[:32000] + "...[truncated]"
            ws_jobs.cell(row=r_idx, column=c_idx, value=value)

    # ── Companies sheet ──
    if companies:
        ws_comp = wb.create_sheet("companies")
        comp_cols = ["id", "title", "url", "source", "_emails", "_telefonos", "_sector", "origin_date"]
        comp_labels = ["id", "company_name", "website", "source", "emails", "telefonos", "sector", "origin_date"]
        for c_idx, label in enumerate(comp_labels, 1):
            cell = ws_comp.cell(row=1, column=c_idx, value=label)
            cell.fill = header_fill
            cell.font = header_font_white
            cell.alignment = Alignment(horizontal="center")
            ws_comp.column_dimensions[get_column_letter(c_idx)].width = max(12, min(40, len(label) + 5))
        ws_comp.freeze_panes = "A2"
        for r_idx, comp in enumerate(companies, 2):
            for c_idx, col in enumerate(comp_cols, 1):
                value = comp.get(col, "")
                ws_comp.cell(row=r_idx, column=c_idx, value=value)

    wb.save(path)


# ── Main ────────────────────────────────────────────────────────────────

def main() -> int:
    print("=" * 60)
    print("LinkedIn Scraper — Data Consolidation")
    print("=" * 60)

    all_records: list[dict] = []

    # 1. Jul 15 Guest API jobs (585 records)
    p = HISTORICAL_DIR / "jobs_20260715.json"
    if p.exists():
        records = _read_guest_api_json(p, "jobs_20260715.json")
        print(f"  jobs_20260715.json: {len(records)} records")
        all_records.extend(records)

    # 2. Jul 23 all_results envelope (80 enriched jobs)
    p = HISTORICAL_DIR / "all_results_20260723_174720.json"
    if p.exists():
        records = _read_all_results_json(p, "all_results_20260723_174720.json")
        print(f"  all_results_20260723: {len(records)} records")
        all_records.extend(records)

    # 3. MCP historical jobs (175 records, different schema)
    p = HISTORICAL_DIR / "linkedin_jobs_mcp_historical.json"
    if p.exists():
        records = _read_mcp_json(p, "linkedin_jobs_mcp_historical.json")
        print(f"  linkedin_jobs_mcp_historical.json: {len(records)} records")
        all_records.extend(records)

    # 4. Empresas (62 company contacts)
    p = ROOT_DIR / "empresas_encontradas.json"
    if p.exists():
        records = _read_empresas_json(p, "empresas_encontradas.json")
        print(f"  empresas_encontradas.json: {len(records)} records")
        all_records.extend(records)

    print(f"\n  Total before dedup: {len(all_records)}")

    # Dedup
    deduped = _dedup(all_records)
    jobs = [r for r in deduped if r["type"] == "job"]
    companies = [r for r in deduped if r["type"] == "company"]
    print(f"  Total after dedup:  {len(deduped)} ({len(jobs)} jobs + {len(companies)} companies)")

    # Create output dir
    OUTPUT_DIR.mkdir(parents=True, exist_ok=True)

    # Write outputs
    json_path = OUTPUT_DIR / "linkedin_all_jobs.json"
    csv_path = OUTPUT_DIR / "linkedin_all_jobs.csv"
    xlsx_path = OUTPUT_DIR / "linkedin_all_jobs.xlsx"

    _write_json(deduped, json_path)
    print(f"\n  JSON -> {json_path}")

    _write_csv(deduped, csv_path)
    print(f"  CSV  -> {csv_path}")

    _write_xlsx(deduped, xlsx_path)
    print(f"  XLSX -> {xlsx_path}")

    print(f"\n{'=' * 60}")
    print(f"SUMMARY")
    print(f"{'=' * 60}")
    print(f"  Jobs:      {len(jobs)}")
    print(f"  Companies: {len(companies)}")
    print(f"  Total:     {len(deduped)}")
    print(f"{'=' * 60}")

    return 0


if __name__ == "__main__":
    sys.exit(main())
